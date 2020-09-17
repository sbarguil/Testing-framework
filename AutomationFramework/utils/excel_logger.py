import openpyxl
import os
from pathlib import Path


class ExcelLogger:
    workbook_path = None
    workbook_name = None
    workbook_columns = None
    workbook = None
    sheet = None
    log_values = {}
    column_indexes = {}

    def __init__(self, workbook_name, columns):
        self.log_values = {}
        self.column_indexes = {}
        not_windows_path = os.path.dirname(os.path.realpath(__file__)).replace('\\', '/')

        self.workbook_path = Path(not_windows_path.replace('/utils', '')) / workbook_name
        self.workbook_name = self.workbook_path.name
        self.workbook_columns = columns

        try:
            self.workbook = openpyxl.load_workbook(self.workbook_path)
            self.sheet = self.workbook.active
        except:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = 'logs'

            for column in range(0, len(self.workbook_columns)):
                self.sheet.cell(row=1, column=column + 1, value=self.workbook_columns[column])

            self.workbook.save(filename=self.workbook_path)

        self.set_column_indexes()
        self.clean_log_values()

    def set_column_indexes(self):
        for column in range(0, len(self.workbook_columns)):
            self.column_indexes[self.sheet.cell(row=1, column=column + 1).value] = column + 1

    def clean_log_values(self):
        for column in self.workbook_columns:
            self.log_values[column] = ''

    def add_value_to_log_column(self, value, column):
        self.log_values[column] = self.log_values[column] + value

    def write_log(self):
        list_row_data = []
        for column in self.workbook_columns:
            list_row_data.append(self.log_values[column])
        row_data = tuple(list_row_data)
        self.sheet.append(row_data)
        self.workbook.save(filename=self.workbook_path)
        self.clean_log_values()
